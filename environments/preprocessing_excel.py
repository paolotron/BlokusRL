import openpyxl as xl
import numpy as np

# original pieces must be defined in the first row of matrix in excel, the script then rotates/flips them

wb = xl.load_workbook('environments\Blokus_Pieces.xlsx', data_only=True) # workbook
ws = wb['Pieces'] # worksheet
d_cell = 11 # dimension of excel cell used
n_pieces = 21 # number of blokus pieces (standard game)
n_variant = 8 # max number of flip and rotation for each piece

for p_id in range(n_pieces):
    print('Piece ID: ' + str(p_id))
    orig_mat = np.zeros((d_cell, d_cell)) # original 11 x 11 matrix
    start_xl_col = 2 + p_id*d_cell # starting column on excel
    start_xl_row = 2 # starting row on excel
    
    # read original piece
    for i in range(d_cell): # rows
        for j in range(d_cell): # columns
            orig_mat[i,j] = ws.cell(row=start_xl_row+i , column=start_xl_col+j).value
            
    # rotate 3 times +90, +180, + 270
    for k in [1, 2, 3]:
        new_mat = np.rot90(orig_mat, k)
        start_xl_row = 2 + k*d_cell
        # write new piece
        for i in range(d_cell): # rows
            for j in range(d_cell): # columns
                ws.cell(row=start_xl_row+i , column=start_xl_col+j).value = new_mat[i,j]
                
    # flip, then rotate 4 times 0, +90, +180, + 270
    flip_mat = np.fliplr(orig_mat) # flip horizontally
    for k in [0, 1, 2, 3]:
        new_mat = np.rot90(flip_mat, k)
        start_xl_row = 2 + (4 + k)*d_cell
        # write new piece
        for i in range(d_cell): # rows
            for j in range(d_cell): # columns
                ws.cell(row=start_xl_row+i , column=start_xl_col+j).value = new_mat[i,j]
    
wb.save('environments\Blokus_Pieces.xlsx')
    
